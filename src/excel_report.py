"""Excel report generation for all pipeline stages (openpyxl only)."""

from __future__ import annotations

import re
from copy import copy as _copy
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import Config
from .io_utils import find_optional_col, safe_output_path
from .od_matrix import build_od_matrices


# ---------------------------------------------------------------------------
# Shared style helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_VCENTER = Alignment(vertical="center")


def _style_header_row(ws, row: int, col_start: int, col_end: int):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _set_borders(ws, r1: int, c1: int, r2: int, c2: int):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _THIN_BORDER


def _write_df(
    ws,
    df: pd.DataFrame,
    start_row: int,
    start_col: int,
    number_formats: Optional[dict] = None,
) -> tuple[int, int]:
    """Write a DataFrame to *ws* with a styled header row.

    Returns (end_row, end_col) of the written block.
    """
    r = start_row
    c = start_col

    # Header
    for j, name in enumerate(df.columns):
        ws.cell(r, c + j, value=str(name))
    _style_header_row(ws, r, c, c + len(df.columns) - 1)
    r += 1

    # Data
    for i in range(len(df)):
        for j, col in enumerate(df.columns):
            ws.cell(r + i, c + j, value=df.iloc[i, j])

    end_row = r + len(df) - 1
    end_col = c + len(df.columns) - 1

    _set_borders(ws, start_row, start_col, end_row, end_col)

    for rr in range(start_row, end_row + 1):
        for cc in range(start_col, end_col + 1):
            ws.cell(rr, cc).alignment = _VCENTER

    if number_formats:
        for col_offset, fmt in number_formats.items():
            col_idx = start_col + col_offset
            for rr in range(start_row + 1, end_row + 1):
                ws.cell(rr, col_idx).number_format = fmt

    return end_row, end_col


def _value_only_labels():
    dl = DataLabelList()
    dl.showVal = True
    dl.showCatName = False
    dl.showSerName = False
    dl.showLegendKey = False
    dl.showPercent = False
    return dl


def _add_bar_chart(
    ws,
    title: str,
    cats_col: int,
    vals_col: int,
    header_row: int,
    last_row: int,
    anchor: str,
    width: int = 18,
    height: int = 8,
):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    data = Reference(ws, min_col=vals_col, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=cats_col, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = _value_only_labels()
    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)


def _add_matrix_chart(
    ws,
    title: str,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
    anchor: str,
):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    data = Reference(ws, min_col=min_col + 1, min_row=min_row, max_col=max_col, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = _value_only_labels()
    chart.height = 10
    chart.width = 22
    ws.add_chart(chart, anchor)


def _bold(ws, row: int, col: int, value):
    cell = ws.cell(row, col, value)
    f = _copy(cell.font)
    f.bold = True
    cell.font = f
    return cell


# ---------------------------------------------------------------------------
# Stage 1 – Filter report
# ---------------------------------------------------------------------------

def write_filter_report(
    cfg: Config,
    df: pd.DataFrame,
    run_tag: str,
) -> Path:
    """Write a minimal filter report: data sheet + basic statistics."""
    out_path = safe_output_path(cfg.output_dir / f"{run_tag}_1_filter.xlsx")

    # Drop geometry columns – Excel can't store Shapely objects
    export_df = df.drop(
        columns=[c for c in df.columns if c.startswith("_geom")],
        errors="ignore",
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Данные", index=False)

        id_col = cfg.columns.id
        total_resp = int(df[id_col].nunique()) if id_col in df.columns else 0
        total_trips = len(df)
        avg = round(total_trips / total_resp, 2) if total_resp else 0

        stats = pd.DataFrame({
            "Показатель": [
                "Количество респондентов",
                "Количество передвижений",
                "Среднее передвижений на чел.",
            ],
            "Значение": [total_resp, total_trips, avg],
        })
        stats.to_excel(writer, sheet_name="Статистика", index=False)

    print(f"  Отчёт этапа 1 сохранён: {out_path.name}")
    return out_path


# ---------------------------------------------------------------------------
# Stage 2 – Status report
# ---------------------------------------------------------------------------

def write_status_report(
    cfg: Config,
    df: pd.DataFrame,
    stats_df: pd.DataFrame,
    run_tag: str,
) -> Path:
    """Write status report: statistics sheet with charts + per-group data sheets."""
    out_path = safe_output_path(cfg.output_dir / f"{run_tag}_2_status.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        stats_df.to_excel(writer, sheet_name="Статистика", index=False)

        for group, group_df in df.groupby("Соцгруппа"):
            sheet_name = str(group)[:31]
            export_df = group_df.drop(
                columns=[c for c in group_df.columns if c.startswith("_geom")],
                errors="ignore",
            )
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Add charts to Statistics sheet
    wb = load_workbook(out_path)
    ws = wb["Статистика"]

    last_row = ws.max_row
    _add_bar_chart(ws, "Количество передвижений", 1, 2, 1, last_row, "G2")
    _add_bar_chart(ws, "Количество респондентов", 1, 3, 1, last_row, "G20")
    _add_bar_chart(ws, "Среднее передвижений на респондента", 1, 4, 1, last_row, "G38")

    out_path = safe_output_path(out_path)  # recheck after load
    wb.save(out_path)
    print(f"  Отчёт этапа 2 сохранён: {out_path.name}")
    return out_path


# ---------------------------------------------------------------------------
# Stage 3 – OD report
# ---------------------------------------------------------------------------

def write_od_report(
    cfg: Config,
    df: pd.DataFrame,
    run_tag: str,
) -> Path:
    """Write OD report: per-social-group sheets with OD tables, matrices, transport."""
    out_path = safe_output_path(cfg.output_dir / f"{run_tag}_3_od.xlsx")

    id_col = cfg.columns.id
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    unmatched_rows = []

    for group, sub_group in df.groupby("Соцгруппа"):
        sheet_name = str(group)[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions["A"].width = 44
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["G"].width = 42

        total_trips = len(sub_group)
        total_resp = int(sub_group[id_col].nunique())

        _bold(ws, 1, 1, f"Статус: {group} — Пары и транспорт")
        ws["A1"].font = Font(size=14, bold=True)

        ws["A3"] = "Количество передвижений"
        ws["B3"] = total_trips
        ws["A4"] = "Количество респондентов (уник. ID)"
        ws["B4"] = total_resp

        # -- OD distribution table --
        ws["A6"] = "Распределение по парам (Группа отправления - Группа прибытия)"
        ws["A6"].font = Font(bold=True)

        od_tab = (
            sub_group["_od_pair"]
            .value_counts()
            .rename_axis("Пара")
            .reset_index(name="Количество")
        )
        od_tab["Доля"] = od_tab["Количество"] / (total_trips or 1)
        od_tab["На 1 респондента"] = od_tab["Количество"] / (total_resp or 1)

        od_start = 7
        od_end_row, _ = _write_df(
            ws,
            od_tab[["Пара", "Количество", "Доля", "На 1 респондента"]],
            od_start, 1,
            number_formats={2: "0.00%", 3: "0.000"},
        )
        _add_bar_chart(ws, "Пары (абс.)", 1, 2, od_start, od_end_row, "G7")

        # -- OD matrices --
        mat_abs, mat_per_resp = build_od_matrices(sub_group, id_col)

        r = od_end_row + 2
        ws[f"A{r}"] = "Матрица O×D (абсолютное количество)"
        ws[f"A{r}"].font = Font(bold=True)
        r += 1

        mat_out = mat_abs.copy()
        mat_out.insert(0, "Откуда\\Куда", mat_out.index)
        mat_out.reset_index(drop=True, inplace=True)
        mat1_start = r
        mat1_end_row, mat1_end_col = _write_df(ws, mat_out, mat1_start, 1)
        _add_matrix_chart(ws, "O×D (абс.)", mat1_start, 1, mat1_end_row, mat1_end_col, f"G{mat1_start}")

        r = mat1_end_row + 2
        ws[f"A{r}"] = "Матрица O×D (кол-во / число респондентов)"
        ws[f"A{r}"].font = Font(bold=True)
        r += 1

        mat2_out = mat_per_resp.copy()
        mat2_out.insert(0, "Откуда\\Куда", mat2_out.index)
        mat2_out.reset_index(drop=True, inplace=True)
        mat2_start = r
        mat2_end_row, mat2_end_col = _write_df(
            ws, mat2_out, mat2_start, 1,
            number_formats={j: "0.000" for j in range(1, len(mat2_out.columns))},
        )
        _add_matrix_chart(
            ws, "O×D (на респондента)", mat2_start, 1, mat2_end_row, mat2_end_col, f"G{mat2_start}"
        )

        # -- Transport overall --
        r = mat2_end_row + 2
        ws[f"A{r}"] = "Распределение по транспорту (группы)"
        ws[f"A{r}"].font = Font(bold=True)
        r += 1

        tr_tab = (
            sub_group["_transport_group"]
            .value_counts()
            .rename_axis("Транспорт (группа)")
            .reset_index(name="Количество")
        )
        tr_tab["Доля"] = tr_tab["Количество"] / (total_trips or 1)
        tr_tab["На 1 респондента"] = tr_tab["Количество"] / (total_resp or 1)

        tr_start = r
        tr_end_row, _ = _write_df(
            ws,
            tr_tab[["Транспорт (группа)", "Количество", "Доля", "На 1 респондента"]],
            tr_start, 1,
            number_formats={2: "0.00%", 3: "0.000"},
        )
        _add_bar_chart(ws, "Транспорт (абс.)", 1, 2, tr_start, tr_end_row, f"G{tr_start}")

        # -- Transport per OD pair --
        r = tr_end_row + 2
        ws[f"A{r}"] = "Транспорт по каждой паре (O-D)"
        ws[f"A{r}"].font = Font(bold=True)
        r += 1

        for od_name in od_tab["Пара"].tolist():
            sub_od = sub_group[sub_group["_od_pair"] == od_name]
            if sub_od.empty:
                continue
            moves_od = len(sub_od)
            sub_tr = (
                sub_od["_transport_group"]
                .value_counts()
                .rename_axis("Транспорт (группа)")
                .reset_index(name="Количество")
            )
            sub_tr["Доля"] = sub_tr["Количество"] / (moves_od or 1)
            sub_tr["На 1 респондента"] = sub_tr["Количество"] / (total_resp or 1)

            ws[f"A{r}"] = f"{od_name} (передвижений: {moves_od})"
            ws[f"A{r}"].font = Font(bold=True)
            r += 1

            blk_start = r
            blk_end_row, _ = _write_df(
                ws,
                sub_tr[["Транспорт (группа)", "Количество", "Доля", "На 1 респондента"]],
                blk_start, 1,
                number_formats={2: "0.00%", 3: "0.000"},
            )
            _add_bar_chart(
                ws, f"Транспорт: {od_name}", 1, 2, blk_start, blk_end_row, f"G{blk_start}"
            )
            r = blk_end_row + 2

        ws.freeze_panes = "A7"

    # Unmatched sheet
    ws_u = wb.create_sheet("НЕ_СОПОСТАВЛЕНО")
    ws_u.column_dimensions["A"].width = 26
    ws_u.column_dimensions["B"].width = 22
    ws_u.column_dimensions["C"].width = 80
    ws_u["A1"] = "Статус"
    ws_u["B1"] = "Поле"
    ws_u["C1"] = "Значение"
    _style_header_row(ws_u, 1, 1, 3)

    out_path_final = safe_output_path(out_path)
    wb.save(out_path_final)
    print(f"  Отчёт этапа 3 сохранён: {out_path_final.name}")
    return out_path_final


# ---------------------------------------------------------------------------
# Stage 4 – EVA report
# ---------------------------------------------------------------------------

def write_eva_report(
    cfg: Config,
    eva_results: dict,
    run_tag: str,
) -> Path:
    """Write EVA report: per-OD-pair sheets with hourly + duration + BW charts."""
    out_path = safe_output_path(cfg.output_dir / f"{run_tag}_4_eva.xlsx")

    bw_cfg = cfg.bw_model
    rows_per_section = bw_cfg.max_minutes // bw_cfg.bin_minutes  # e.g. 24

    # -- Summary sheet first --
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_rows = [
            {"Пара": od_pair, "Передвижения": len(data["sub_df"])}
            for od_pair, data in eva_results.items()
        ]
        summary_df = pd.DataFrame(summary_rows).sort_values("Передвижения", ascending=False)
        summary_df.to_excel(writer, sheet_name="Сводка", index=False)

        for od_pair, data in eva_results.items():
            sheet = str(od_pair)[:31]
            hours_df = data["hours_df"]
            hours_df.to_excel(writer, sheet_name=sheet, index=False, startrow=0)

            current_row = len(hours_df) + 4  # gap after hours table

            for tg, tg_data in data["transports"].items():
                dur_df = tg_data["dur_df"]

                # Section header
                pd.DataFrame({
                    "": [f"Длительность + BW: транспортная группа = {tg}"]
                }).to_excel(
                    writer, sheet_name=sheet, index=False, header=False, startrow=current_row
                )
                current_row += 2

                dur_df.to_excel(writer, sheet_name=sheet, index=False, startrow=current_row)

                # BW params table to the right (column J = 10, index 9)
                params_df = pd.DataFrame({
                    "Параметр": ["E", "F", "G", "MAE", "RMSE", "R2"],
                    "Значение": [
                        tg_data["E"],
                        tg_data["F"],
                        tg_data["G"],
                        tg_data["metrics"]["MAE"],
                        tg_data["metrics"]["RMSE"],
                        tg_data["metrics"]["R2"],
                    ],
                })
                params_df.to_excel(
                    writer, sheet_name=sheet, index=False,
                    startrow=current_row, startcol=9,
                )

                current_row += rows_per_section + 4

    # -- Second pass: add charts and Excel BW formulas --
    wb = load_workbook(out_path)

    for sh in wb.sheetnames:
        if sh == "Сводка":
            continue
        ws = wb[sh]

        # Hourly chart
        chart_h = BarChart()
        chart_h.type = "col"
        chart_h.grouping = "clustered"
        chart_h.title = "Распределение по часам"
        chart_h.x_axis.title = "Час суток"
        chart_h.y_axis.title = "Передвижения"
        data_h = Reference(ws, min_col=2, min_row=1, max_row=25)
        cats_h = Reference(ws, min_col=1, min_row=2, max_row=25)
        chart_h.add_data(data_h, titles_from_data=True)
        chart_h.set_categories(cats_h)
        chart_h.dataLabels = _value_only_labels()
        chart_h.height = 9
        chart_h.width = 22
        ws.add_chart(chart_h, "E2")

        # Find each duration section and add formulas + charts
        row = 1
        section_idx = 0
        while row <= ws.max_row:
            if ws.cell(row, 1).value == "Интервал (мин)":
                section_idx += 1
                dur_header = row
                dur_first = dur_header + 1
                dur_last = dur_header + rows_per_section

                # BW param cells (column K = 11)
                row_E = dur_header + 1
                row_F = dur_header + 2
                row_G = dur_header + 3
                cell_E = f"$K${row_E}"
                cell_F = f"$K${row_F}"
                cell_G = f"$K${row_G}"

                # Write BW formulas for columns F (6) and G (7)
                for rr in range(dur_first, dur_last + 1):
                    W_cell = ws.cell(rr, 2).coordinate   # W (верх)
                    diff_cell = ws.cell(rr, 5).coordinate  # Разница с предыдущим
                    bw_cell = ws.cell(rr, 6).coordinate    # BW (модель)
                    err_cell = ws.cell(rr, 7).coordinate   # |ошибка|
                    ws[bw_cell].value = (
                        f"=1-1/POWER(1+{W_cell},"
                        f"{cell_E}/(1+EXP({cell_F}-{cell_G}*{W_cell})))"
                    )
                    ws[err_cell].value = f"=ABS({bw_cell}-{diff_cell})"

                # Duration bar chart (absolute counts)
                chart_d = BarChart()
                chart_d.type = "col"
                chart_d.grouping = "clustered"
                chart_d.title = f"Длительность (абс.) — секция {section_idx}"
                chart_d.x_axis.title = "Интервал, мин"
                chart_d.y_axis.title = "Передвижения"
                data_d = Reference(ws, min_col=3, min_row=dur_header, max_row=dur_last)
                cats_d = Reference(ws, min_col=1, min_row=dur_first, max_row=dur_last)
                chart_d.add_data(data_d, titles_from_data=True)
                chart_d.set_categories(cats_d)
                chart_d.dataLabels = _value_only_labels()
                chart_d.height = 8
                chart_d.width = 20
                ws.add_chart(chart_d, f"E{dur_header}")

                # Comparison: Разница vs BW (line chart)
                chart_cmp = LineChart()
                chart_cmp.title = f"Разница vs BW — секция {section_idx}"
                chart_cmp.x_axis.title = "W (верх), мин"
                chart_cmp.y_axis.title = "Значение"
                cats_cmp = Reference(ws, min_col=2, min_row=dur_first, max_row=dur_last)
                ser_actual = Reference(ws, min_col=5, min_row=dur_header, max_row=dur_last)
                ser_model = Reference(ws, min_col=6, min_row=dur_header, max_row=dur_last)
                chart_cmp.add_data(ser_actual, titles_from_data=True)
                chart_cmp.add_data(ser_model, titles_from_data=True)
                chart_cmp.set_categories(cats_cmp)
                chart_cmp.dataLabels = _value_only_labels()
                chart_cmp.height = 8
                chart_cmp.width = 26
                ws.add_chart(chart_cmp, f"M{dur_header}")

                row = dur_last + 1
                continue
            row += 1

        # Statistics block far to the right (column AD = 30)
        _write_stats_block(
            ws,
            start_row=1,
            start_col=30,
            sub_df=eva_results.get(sh, {}).get("sub_df", pd.DataFrame()),
            transport_col="_transport_group",
            cfg=cfg,
        )
        for col_idx in range(30, 38):
            ws.column_dimensions[get_column_letter(col_idx)].width = 42

    out_path_final = safe_output_path(out_path)
    wb.save(out_path_final)
    print(f"  Отчёт этапа 4 сохранён: {out_path_final.name}")
    return out_path_final


# ---------------------------------------------------------------------------
# Statistics block helper (used in EVA report)
# ---------------------------------------------------------------------------

def _write_stats_block(
    ws,
    start_row: int,
    start_col: int,
    sub_df: pd.DataFrame,
    transport_col: str,
    cfg: Config,
):
    """Write a statistics block to the right of the main data area."""
    if sub_df.empty:
        return

    r = start_row
    c = start_col

    def write(text, row=None, col=None, bold=False):
        nonlocal r
        if row is None:
            row = r
        if col is None:
            col = c
        cell = ws.cell(row, col, text)
        if bold:
            f = _copy(cell.font)
            f.bold = True
            cell.font = f
        return row, col

    def next_row(n=1):
        nonlocal r
        r += n

    write("СТАТИСТИКА ПО ЛИСТУ", bold=True)
    next_row(2)

    # Fields split by transport group
    for field in cfg.stat_fields_by_transport:
        col_name = find_optional_col(sub_df, field)
        if col_name is None:
            write(f"{field}: (нет столбца)")
            next_row(2)
            continue

        write(field, bold=True)
        next_row()

        header_row = r
        for j, h in enumerate(["Транспортная группа", "N", "Среднее", "Медиана", "Мин", "Макс"]):
            write(h, row=header_row, col=c + j, bold=True)
        r = header_row + 1
        start_data = r
        tg_count = 0

        if transport_col in sub_df.columns:
            for tg, tg_df in sub_df.groupby(transport_col):
                ser = tg_df[col_name].dropna()
                ser = ser[ser.astype(str).str.strip().ne("")]
                n = len(ser)
                write(str(tg), row=r, col=c)
                write(int(n), row=r, col=c + 1)
                num = pd.to_numeric(ser, errors="coerce").dropna()
                if len(num) >= max(3, int(0.5 * max(1, n))):
                    write(round(float(num.mean()), 2), row=r, col=c + 2)
                    write(round(float(num.median()), 2), row=r, col=c + 3)
                    write(round(float(num.min()), 2), row=r, col=c + 4)
                    write(round(float(num.max()), 2), row=r, col=c + 5)
                r += 1
                tg_count += 1

        end_data = start_data + tg_count - 1
        if tg_count > 0:
            try:
                mean_ref = Reference(ws, min_col=c + 2, min_row=start_data, max_row=end_data)
                cats_ref = Reference(ws, min_col=c, min_row=start_data, max_row=end_data)
                chart = BarChart()
                chart.type = "col"
                chart.title = f"{field[:30]}: среднее"
                chart.add_data(mean_ref)
                chart.set_categories(cats_ref)
                chart.dataLabels = _value_only_labels()
                chart.height = 8
                chart.width = 20
                ws.add_chart(chart, f"{get_column_letter(c + 7)}{header_row}")
            except Exception:
                pass

        next_row(2)

    # Simple aggregate fields
    for field in cfg.stat_fields_simple:
        col_name = find_optional_col(sub_df, field)
        if col_name is None:
            write(f"{field}: (нет столбца)")
            next_row(2)
            continue

        write(field, bold=True)
        next_row()

        ser = sub_df[col_name].dropna()
        ser = ser[ser.astype(str).str.strip().ne("")]
        n_total = len(sub_df[col_name])
        n_non = len(ser)

        write("Всего", row=r, col=c, bold=True); write(int(n_total), row=r, col=c + 1); r += 1
        write("Заполнено", row=r, col=c, bold=True); write(int(n_non), row=r, col=c + 1); r += 1
        write("Пусто", row=r, col=c, bold=True); write(int(n_total - n_non), row=r, col=c + 1); r += 1

        num = pd.to_numeric(ser, errors="coerce").dropna()
        if len(num) >= max(5, int(0.5 * max(1, n_non))):
            write("Среднее", row=r, col=c, bold=True); write(round(float(num.mean()), 2), row=r, col=c + 1); r += 1
            write("Медиана", row=r, col=c, bold=True); write(round(float(num.median()), 2), row=r, col=c + 1); r += 1
            write("Мин", row=r, col=c, bold=True); write(round(float(num.min()), 2), row=r, col=c + 1); r += 1
            write("Макс", row=r, col=c, bold=True); write(round(float(num.max()), 2), row=r, col=c + 1); r += 1

        next_row(2)
