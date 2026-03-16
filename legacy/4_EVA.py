
import re
import numpy as np
import pandas as pd
from pathlib import Path
from collections import defaultdict, Counter
from copy import copy as _copy

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


# ================== НАСТРОЙКИ ==================
INPUT_WB = None                 # выбирается пользователем
PAIRS_XLSX = "Пары.xlsx"        # группы пунктов отправления/прибытия
TRANSPORT_XLSX = "Транспорт.xlsx"

OUTPUT_WB = None                # формируется из имени входного файла

ORIG_COL = "Пункт отправления"
DEST_COL = "Пункт прибытия"
TIME_DEP_COL = "Время отправления"
TIME_ARR_COL = "Время прибытия"
TRANSPORT_COL = "Транспорт"
COMMENT_COL = "Комментарий"

BIN_MINUTES = 5
MAX_MINUTES = 120

# границы поиска параметров BW
E_BOUNDS = (0.01, 20.0)
F_BOUNDS = (-20.0, 20.0)
G_BOUNDS = (-2.0, 2.0)

# Статистика в Excel начиная с колонки AD
STATS_START_COL = 30  # AD

# ПОЛЯ ДЛЯ СТАТИСТИКИ (будут искаться startswith)
STAT_FIELDS_SIMPLE = [
    "Количество пересадок",
    "Время ожидания между пересадками",
]

# Эти поля считаем по транспортным группам (как просили)
STAT_FIELDS_BY_TRANSPORT = [
    "Количество людей в машине",
    "Пешеходный подход к начальной остановке или парковке",
    "Время ожидания транспорта",
    "Пешеходный подход от конечной остановки или парковки",
    "Стоимость поездки или парковки",
]
# ==============================================


def normalize_excel_text(x):
    if pd.isna(x):
        return None
    s = str(x)
    s = s.replace("\u00A0", " ").replace("\u2007", " ").replace("\u202F", " ")
    s = s.replace("\r", " ").replace("\n", " ")
    s = re.sub(r"[ \t]+", " ", s).strip()
    return s.lower() if s else None


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.replace("\r", "", regex=False)
        .str.replace("\n", " ", regex=False)
        .str.strip()
    )
    return df


def choose_input_xlsx(exclude_names=None) -> str:
    exclude_names = set(exclude_names or [])
    files = sorted(
        [p for p in Path(".").iterdir()
         if p.is_file() and p.suffix.lower() == ".xlsx" and p.name not in exclude_names]
    )
    if not files:
        raise FileNotFoundError("В текущем каталоге нет файлов .xlsx для выбора.")

    print("Доступные файлы Excel (.xlsx):")
    for i, f in enumerate(files, start=1):
        print(f"{i}. {f.name}")

    while True:
        raw = input("Выберите номер файла: ").strip()
        try:
            idx = int(raw)
            if 1 <= idx <= len(files):
                return files[idx - 1].name
        except ValueError:
            pass
        print("❌ Неверный ввод. Введите число из списка.")

def ensure_writable_output_path(path_str: str) -> str:
    """
    Если файл назначения занят (открыт в Excel) или недоступен для записи,
    подбираем новое имя: <stem>_1.xlsx, <stem>_2.xlsx, ...
    Возвращает путь, который можно создать/перезаписать.
    """
    p = Path(path_str)
    # если папка недоступна — пусть упадёт дальше с понятной ошибкой
    parent = p.parent if str(p.parent) else Path(".")
    stem = p.stem
    suffix = p.suffix or ".xlsx"

    # быстрый тест записи: пытаемся открыть файл в режиме append (или создать)
    def _can_open(candidate: Path) -> bool:
        try:
            # 'a' создаёт файл если нет и проверяет право записи
            with open(candidate, "a+b"):
                return True
        except PermissionError:
            return False
        except OSError:
            # например, путь некорректен
            return False

    cand = p
    if _can_open(cand):
        return str(cand)

    for i in range(1, 200):
        cand = parent / f"{stem}_{i}{suffix}"
        if _can_open(cand):
            return str(cand)

    raise PermissionError(f"Не удалось подобрать доступное имя для файла: {path_str}")



def parse_columns_as_groups(path_xlsx: str):
    """
    Пары.xlsx:
      - заголовки столбцов = названия групп
      - ниже значения = варианты пунктов
    """
    df = pd.read_excel(path_xlsx, header=0, dtype=str)
    df = df.loc[:, ~df.columns.astype(str).str.lower().str.startswith("unnamed")]

    mapping = {}
    conflicts = defaultdict(list)

    for col in df.columns:
        group_name = str(col).strip()
        if not group_name:
            continue
        for raw_val in df[col].dropna().tolist():
            key = normalize_excel_text(raw_val)
            if not key:
                continue
            if key in mapping and mapping[key] != group_name:
                if mapping[key] not in conflicts[key]:
                    conflicts[key].append(mapping[key])
                if group_name not in conflicts[key]:
                    conflicts[key].append(group_name)
                continue
            mapping[key] = group_name

    return mapping, conflicts


def parse_transport_groups(path_xlsx: str):
    """
    Транспорт.xlsx:
      - заголовки столбцов = названия транспортных групп
      - ниже = варианты значений из поля "Транспорт"
    """
    df = pd.read_excel(path_xlsx, header=0, dtype=str)
    df = df.loc[:, ~df.columns.astype(str).str.lower().str.startswith("unnamed")]

    mapping = {}
    for col in df.columns:
        group = str(col).strip()
        if not group:
            continue
        for raw in df[col].dropna().tolist():
            key = normalize_excel_text(raw)
            if not key:
                continue
            mapping[key] = group

    group_names = [str(c).strip() for c in df.columns if str(c).strip()]
    return mapping, group_names


def load_all_sheets(path_xlsx: str) -> pd.DataFrame:
    xls = pd.ExcelFile(path_xlsx)
    frames = []
    for sh in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sh)
        df = normalize_columns(df)
        df["__sheet"] = sh
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def find_col(df: pd.DataFrame, target: str) -> str:
    if target in df.columns:
        return target
    cand = [c for c in df.columns if str(c).strip().startswith(target)]
    if not cand:
        raise KeyError(f"Не найден столбец '{target}'. Доступные: {list(df.columns)}")
    return cand[0]


def find_optional_col(df: pd.DataFrame, target: str):
    if target in df.columns:
        return target
    cand = [c for c in df.columns if str(c).strip().startswith(target)]
    return cand[0] if cand else None


def extract_hour(val):
    if pd.isna(val):
        return None

    if hasattr(val, "hour"):
        try:
            return int(val.hour)
        except Exception:
            pass

    if isinstance(val, (int, float)):
        x = float(val)
        if 0 <= x < 1:
            return int(x * 24)
        dt = pd.to_datetime(x, unit="D", origin="1899-12-30", errors="coerce")
        return int(dt.hour) if pd.notna(dt) else None

    s = str(val).strip()
    if not s:
        return None

    dt = pd.to_datetime(s, errors="coerce")
    if pd.notna(dt):
        return int(dt.hour)

    m = re.match(r"^\s*(\d{1,2})\s*[:.]\s*(\d{1,2})", s)
    if m:
        h = int(m.group(1))
        return h if 0 <= h <= 23 else None

    return None


def time_to_minutes(val):
    if pd.isna(val):
        return None

    if hasattr(val, "hour") and hasattr(val, "minute"):
        try:
            return int(val.hour) * 60 + int(val.minute)
        except Exception:
            pass

    if isinstance(val, (int, float)):
        x = float(val)
        if 0 <= x < 1:
            mins = int(round(x * 24 * 60))
            return max(0, min(1439, mins))
        dt = pd.to_datetime(x, unit="D", origin="1899-12-30", errors="coerce")
        if pd.notna(dt):
            return int(dt.hour) * 60 + int(dt.minute)
        return None

    s = str(val).strip()
    if not s:
        return None

    dt = pd.to_datetime(s, errors="coerce")
    if pd.notna(dt):
        return int(dt.hour) * 60 + int(dt.minute)

    m = re.match(r"^\s*(\d{1,2})\s*[:.]\s*(\d{1,2})", s)
    if m:
        h = int(m.group(1))
        mi = int(m.group(2))
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return h * 60 + mi

    return None


def apply_value_only_datalabels(chart):
    dl = DataLabelList()
    dl.showVal = True
    dl.showCatName = True   # подпись интервалов / часов снизу
    dl.showSerName = False
    dl.showLegendKey = False
    dl.showPercent = False
    chart.dataLabels = dl



def build_duration_bins(bin_minutes: int, max_minutes: int):
    if max_minutes % bin_minutes != 0:
        raise ValueError("MAX_MINUTES должен делиться на BIN_MINUTES без остатка.")
    bins = list(range(0, max_minutes + bin_minutes, bin_minutes))  # 0..120
    labels = [f"{bins[i]}-{bins[i+1]}" for i in range(len(bins) - 1)]
    w_upper = [bins[i+1] for i in range(len(bins) - 1)]
    return bins, labels, w_upper


# ======= МОДЕЛЬ BW =======
def phi(W, E, F, G):
    return E / (1.0 + np.exp(F - G * W))


def bw(W, E, F, G):
    return 1.0 - 1.0 / np.power(1.0 + W, phi(W, E, F, G))


def fit_bw_params(W, y):
    W = np.asarray(W, dtype=float)
    y = np.asarray(y, dtype=float)

    mask = np.isfinite(W) & np.isfinite(y)
    W = W[mask]
    y = y[mask]

    if len(W) < 6:
        return 2.0, 0.0, 0.1

    try:
        from scipy.optimize import curve_fit  # type: ignore

        def f(W_, E_, F_, G_):
            return bw(W_, E_, F_, G_)

        p0 = [2.0, 0.0, 0.05]
        bounds = ([E_BOUNDS[0], F_BOUNDS[0], G_BOUNDS[0]],
                  [E_BOUNDS[1], F_BOUNDS[1], G_BOUNDS[1]])
        popt, _ = curve_fit(f, W, y, p0=p0, bounds=bounds, maxfev=50000)
        return float(popt[0]), float(popt[1]), float(popt[2])
    except Exception:
        pass

    rng = np.random.default_rng(42)

    def sse(params):
        E_, F_, G_ = params
        pred = bw(W, E_, F_, G_)
        return float(np.sum((pred - y) ** 2))

    best = None
    best_sse = float("inf")

    for _ in range(6000):
        E_ = rng.uniform(*E_BOUNDS)
        F_ = rng.uniform(*F_BOUNDS)
        G_ = rng.uniform(*G_BOUNDS)
        val = sse((E_, F_, G_))
        if val < best_sse:
            best_sse = val
            best = (E_, F_, G_)

    E_, F_, G_ = best
    step_E = (E_BOUNDS[1] - E_BOUNDS[0]) / 40
    step_F = (F_BOUNDS[1] - F_BOUNDS[0]) / 40
    step_G = (G_BOUNDS[1] - G_BOUNDS[0]) / 40

    for _ in range(2000):
        cand = (
            float(np.clip(rng.normal(E_, step_E), *E_BOUNDS)),
            float(np.clip(rng.normal(F_, step_F), *F_BOUNDS)),
            float(np.clip(rng.normal(G_, step_G), *G_BOUNDS)),
        )
        val = sse(cand)
        if val < best_sse:
            best_sse = val
            E_, F_, G_ = cand
            step_E *= 0.995
            step_F *= 0.995
            step_G *= 0.995

    return float(E_), float(F_), float(G_)


def calc_stats(y_true, y_pred):
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)

    mask = np.isfinite(y_true) & np.isfinite(y_pred)
    y_true = y_true[mask]
    y_pred = y_pred[mask]

    if len(y_true) == 0:
        return {"MAE": None, "RMSE": None, "R2": None}

    mae = float(np.mean(np.abs(y_true - y_pred)))
    rmse = float(np.sqrt(np.mean((y_true - y_pred) ** 2)))
    ss_res = float(np.sum((y_true - y_pred) ** 2))
    ss_tot = float(np.sum((y_true - np.mean(y_true)) ** 2))
    r2 = float(1 - ss_res / ss_tot) if ss_tot > 0 else None
    return {"MAE": mae, "RMSE": rmse, "R2": r2}


def first_transport_token(val: str):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    parts = re.split(r"[;,/|]+", s)
    token = parts[0].strip()
    return token if token else None


def summarize_comments(series: pd.Series, top_n=8):
    texts = [str(x).strip() for x in series.dropna().tolist() if str(x).strip()]
    total = len(texts)
    if total == 0:
        return {"total": 0, "summary_lines": ["Нет комментариев."]}

    stop = set([
        "и","в","во","на","не","что","это","а","я","мы","вы","он","она","они",
        "по","к","ко","с","со","из","за","для","или","же","как","так","то","да",
        "нет","бы","же","там","тут","очень","просто","у","от","до","при","ли"
    ])
    words = []
    for t in texts:
        t2 = re.sub(r"[^0-9a-zA-Zа-яА-ЯёЁ\s-]", " ", t.lower())
        for w in t2.split():
            if len(w) < 4 or w in stop:
                continue
            words.append(w)

    top_words = Counter(words).most_common(top_n)
    top_phrases = Counter([t[:80] for t in texts]).most_common(5)

    lines = [
        f"Комментариев (непустых): {total}",
        "Топ ключевых слов: " + ", ".join([f"{w} ({c})" for w, c in top_words]) if top_words else "Топ ключевых слов: —",
        "Частые формулировки: " + " | ".join([f"«{p}» ({c})" for p, c in top_phrases]) if top_phrases else "Частые формулировки: —",
    ]
    return {"total": total, "summary_lines": lines}


def _write_bold(ws, row, col, value):
    cell = ws.cell(row, col, value)
    f = _copy(cell.font)
    f.bold = True
    cell.font = f
    return cell


def write_stats_block(ws, start_row: int, start_col: int, sub_df: pd.DataFrame, transport_group_col: str):
    """
    Блок статистики справа от AD.
    Значения — в отдельных ячейках, по возможности — графики.
    """
    r = start_row
    c = start_col

    def write(text, row=None, col=None, bold=False):
        nonlocal r
        if row is None:
            row = r
        if col is None:
            col = c
        if bold:
            _write_bold(ws, row, col, text)
        else:
            ws.cell(row, col, text)
        return row, col

    def next_row(n=1):
        nonlocal r
        r += n

    write("СТАТИСТИКА ПО ЛИСТУ", bold=True)
    next_row(2)

    # 1) По транспортным группам
    for field in STAT_FIELDS_BY_TRANSPORT:
        col_name = find_optional_col(sub_df, field)
        if col_name is None:
            write(f"{field}: (нет столбца)")
            next_row(2)
            continue

        write(field, bold=True)
        next_row(1)

        header_row = r
        headers = ["Транспортная группа", "N", "Среднее", "Медиана", "Мин", "Макс"]
        for j, h in enumerate(headers):
            write(h, row=header_row, col=c + j, bold=True)
        r = header_row + 1

        start_data_row = r
        tg_rows = 0

        for tg, tg_df in sub_df.groupby(transport_group_col):
            ser = tg_df[col_name]
            non_empty = ser.dropna()
            non_empty = non_empty[non_empty.astype(str).str.strip().ne("")]
            n = len(non_empty)

            write(str(tg), row=r, col=c + 0)
            write(int(n), row=r, col=c + 1)

            num = pd.to_numeric(non_empty, errors="coerce")
            if num.notna().sum() >= max(3, int(0.5 * max(1, n))):
                num = num.dropna()
                write(float(num.mean()), row=r, col=c + 2)
                write(float(num.median()), row=r, col=c + 3)
                write(float(num.min()), row=r, col=c + 4)
                write(float(num.max()), row=r, col=c + 5)
            else:
                vc = non_empty.astype(str).str.strip().value_counts()
                top1 = vc.index[0] if not vc.empty else ""
                write(str(top1), row=r, col=c + 2)

            r += 1
            tg_rows += 1

        end_data_row = start_data_row + tg_rows - 1

        # график по среднему (если есть)
        try:
            mean_ref = Reference(ws, min_col=c + 2, min_row=start_data_row, max_row=end_data_row)
            cats_ref = Reference(ws, min_col=c + 0, min_row=start_data_row, max_row=end_data_row)
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = f"{field}: среднее по транспорту"
            chart.x_axis.title = "Транспортная группа"
            chart.y_axis.title = field
            chart.add_data(mean_ref, titles_from_data=False)
            chart.set_categories(cats_ref)
            apply_value_only_datalabels(chart)
            chart.height = 8
            chart.width = 20
            ws.add_chart(chart, f"{get_column_letter(c + 7)}{header_row}")
        except Exception:
            pass

        next_row(2)

    # 2) По листу (без разбивки)
    for field in STAT_FIELDS_SIMPLE:
        col_name = find_optional_col(sub_df, field)
        if col_name is None:
            write(f"{field}: (нет столбца)")
            next_row(2)
            continue

        write(field, bold=True)
        next_row(1)

        ser = sub_df[col_name]
        non_empty = ser.dropna()
        non_empty = non_empty[non_empty.astype(str).str.strip().ne("")]
        n_total = len(ser)
        n_non = len(non_empty)

        write("Всего", row=r, col=c, bold=True); write(int(n_total), row=r, col=c+1); r += 1
        write("Заполнено", row=r, col=c, bold=True); write(int(n_non), row=r, col=c+1); r += 1
        write("Пусто", row=r, col=c, bold=True); write(int(n_total - n_non), row=r, col=c+1); r += 1

        num = pd.to_numeric(non_empty, errors="coerce")
        if num.notna().sum() >= max(5, int(0.5 * max(1, n_non))):
            num = num.dropna()
            write("Среднее", row=r, col=c, bold=True); write(float(num.mean()), row=r, col=c+1); r += 1
            write("Медиана", row=r, col=c, bold=True); write(float(num.median()), row=r, col=c+1); r += 1
            write("Мин", row=r, col=c, bold=True); write(float(num.min()), row=r, col=c+1); r += 1
            write("Макс", row=r, col=c, bold=True); write(float(num.max()), row=r, col=c+1); r += 1

            # мини-график mean/median
            try:
                label_row = r
                write("Среднее", row=label_row, col=c+3)
                write("Медиана", row=label_row+1, col=c+3)
                write(float(num.mean()), row=label_row, col=c+4)
                write(float(num.median()), row=label_row+1, col=c+4)
                data_ref = Reference(ws, min_col=c+4, min_row=label_row, max_row=label_row+1)
                cats_ref = Reference(ws, min_col=c+3, min_row=label_row, max_row=label_row+1)
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "clustered"
                chart.title = f"{field}: mean/median"
                chart.y_axis.title = field
                chart.add_data(data_ref, titles_from_data=False)
                chart.set_categories(cats_ref)
                apply_value_only_datalabels(chart)
                chart.height = 6
                chart.width = 14
                ws.add_chart(chart, f"{get_column_letter(c + 7)}{label_row}")
                r = label_row + 3
            except Exception:
                pass
        else:
            vc = non_empty.astype(str).str.strip().value_counts().head(10)
            write("Значение", row=r, col=c, bold=True); write("Кол-во", row=r, col=c+1, bold=True); r += 1
            start_v = r
            for k, v in vc.items():
                write(str(k), row=r, col=c)
                write(int(v), row=r, col=c+1)
                r += 1
            end_v = r - 1
            try:
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "clustered"
                chart.title = f"{field}: топ значений"
                chart.y_axis.title = "Кол-во"
                data_ref = Reference(ws, min_col=c+1, min_row=start_v, max_row=end_v)
                cats_ref = Reference(ws, min_col=c, min_row=start_v, max_row=end_v)
                chart.add_data(data_ref, titles_from_data=False)
                chart.set_categories(cats_ref)
                apply_value_only_datalabels(chart)
                chart.height = 7
                chart.width = 20
                ws.add_chart(chart, f"{get_column_letter(c + 7)}{start_v}")
            except Exception:
                pass

        next_row(2)

    # 3) Комментарии
    write("Комментарий (обобщение)", bold=True)
    next_row(1)
    col_c = find_optional_col(sub_df, COMMENT_COL)
    if col_c is None:
        write("(нет столбца)")
    else:
        digest = summarize_comments(sub_df[col_c])
        for line in digest["summary_lines"]:
            write(line)
            next_row(1)

    for col in range(start_col, start_col + 12):
        ws.column_dimensions[get_column_letter(col)].width = 28


def main():
    global INPUT_WB, OUTPUT_WB

    INPUT_WB = choose_input_xlsx(exclude_names={PAIRS_XLSX, TRANSPORT_XLSX})
    stem = Path(INPUT_WB).stem
    OUTPUT_WB = f"{stem}_по_парам_час_и_длительность_по_транспорту.xlsx"
    print(f"\nИспользуется входной файл: {INPUT_WB}")
    print(f"Выходной файл: {OUTPUT_WB}\n")

    for p in [INPUT_WB, PAIRS_XLSX, TRANSPORT_XLSX]:
        if not Path(p).exists():
            raise FileNotFoundError(f"Не найден файл: {p}")

    pairs_map, pairs_conflicts = parse_columns_as_groups(PAIRS_XLSX)
    if pairs_conflicts:
        print("⚠ ВНИМАНИЕ: в Пары.xlsx есть конфликты (одно значение в нескольких группах). Примеры:")
        for k, gs in list(pairs_conflicts.items())[:20]:
            print(f"  {k!r} -> {gs}")

    transport_map, _ = parse_transport_groups(TRANSPORT_XLSX)

    df = load_all_sheets(INPUT_WB)
    if df.empty:
        raise ValueError("Входной файл пустой (не найдено данных).")

    orig_col = find_col(df, ORIG_COL)
    dest_col = find_col(df, DEST_COL)
    dep_col = find_col(df, TIME_DEP_COL)
    arr_col = find_col(df, TIME_ARR_COL)
    tr_col = find_col(df, TRANSPORT_COL)

    # --- группы пунктов и пары ---
    df["_o_key"] = df[orig_col].apply(normalize_excel_text)
    df["_d_key"] = df[dest_col].apply(normalize_excel_text)
    df["_o_grp"] = df["_o_key"].map(pairs_map).fillna("Не определено")
    df["_d_grp"] = df["_d_key"].map(pairs_map).fillna("Не определено")
    df["_od_pair"] = df["_o_grp"].astype(str) + " - " + df["_d_grp"].astype(str)

    # --- транспортная группа ---
    def map_transport_group(x):
        if pd.isna(x) or str(x).strip() == "":
            return "Пешком"
        token = first_transport_token(str(x))
        if token is None:
            return "Пешком"
        key = normalize_excel_text(token)
        return transport_map.get(key, "Не определено")

    df["_transport_group"] = df[tr_col].apply(map_transport_group)

    # --- час отправления ---
    df["_hour"] = df[dep_col].apply(extract_hour)

    # --- длительность (NaN-safe) ---
    dep_m = df[dep_col].apply(time_to_minutes)
    arr_m = df[arr_col].apply(time_to_minutes)
    durs = []
    for d, a in zip(dep_m, arr_m):
        # d/a могут быть None или NaN
        if d is None or a is None or (isinstance(d, float) and np.isnan(d)) or (isinstance(a, float) and np.isnan(a)):
            durs.append(None)
            continue
        x = int(a) - int(d)
        if x < 0:
            x += 24 * 60
        durs.append(x)
    df["_dur_min"] = pd.to_numeric(pd.Series(durs), errors="coerce")

    bins, labels, w_upper = build_duration_bins(BIN_MINUTES, MAX_MINUTES)

    # ========== запись в Excel ==========
    OUTPUT_WB = ensure_writable_output_path(OUTPUT_WB)
    with pd.ExcelWriter(OUTPUT_WB, engine="openpyxl") as writer:
        summary = (
            df.groupby("_od_pair")
              .size()
              .rename("Передвижения")
              .reset_index()
              .sort_values("Передвижения", ascending=False)
        )
        summary.rename(columns={"_od_pair": "Пара"}, inplace=True)
        summary.to_excel(writer, sheet_name="Сводка", index=False)

        for od_pair, sub_pair in df.groupby("_od_pair"):
            sheet = str(od_pair)[:31]

            # 1) распределение по часам (общее для пары)
            sub_h = sub_pair[sub_pair["_hour"].notna()].copy()
            sub_h["_hour"] = pd.to_numeric(sub_h["_hour"], errors="coerce")
            sub_h = sub_h[sub_h["_hour"].notna()].copy()
            sub_h["_hour"] = sub_h["_hour"].astype(int)

            counts_h = sub_h["_hour"].value_counts().reindex(range(24), fill_value=0).sort_index()
            hours_df = pd.DataFrame({"Час": counts_h.index, "Передвижения": counts_h.values})
            total_h = int(hours_df["Передвижения"].sum())
            hours_df["Доля"] = (hours_df["Передвижения"] / (total_h if total_h else 1)).round(4)
            hours_df.to_excel(writer, sheet_name=sheet, index=False, startrow=0)

            # 2) Далее, после часов, делаем распределение по длительности ДЛЯ КАЖДОЙ транспортной группы
            current_row = len(hours_df) + 4  # место под заголовок/отступ

            for tg, sub_tg in sub_pair.groupby("_transport_group"):
                # Заголовок секции
                pd.DataFrame({"": [f"Длительность + BW: транспортная группа = {tg}"]}).to_excel(
                    writer, sheet_name=sheet, index=False, header=False, startrow=current_row
                )
                current_row += 2

                sub_d = sub_tg[sub_tg["_dur_min"].notna()].copy()
                sub_d["_dur_min"] = pd.to_numeric(sub_d["_dur_min"], errors="coerce")
                sub_d = sub_d[sub_d["_dur_min"].notna()].copy()
                sub_d = sub_d[(sub_d["_dur_min"] >= 0) & (sub_d["_dur_min"] <= MAX_MINUTES)].copy()

                # 120 -> чуть меньше для right=False
                sub_d["_dur_adj"] = sub_d["_dur_min"].where(sub_d["_dur_min"] < MAX_MINUTES, MAX_MINUTES - 1e-6)

                sub_d["_bin"] = pd.cut(
                    sub_d["_dur_adj"], bins=bins, right=False, labels=labels, include_lowest=True
                )
                counts_d = sub_d["_bin"].value_counts().reindex(labels, fill_value=0).astype(int)

                dur_df = pd.DataFrame({
                    "Интервал (мин)": labels,
                    "W (верх, мин)": w_upper,
                    "Количество": counts_d.values
                })
                total_d = int(dur_df["Количество"].sum())
                dur_df["Доля"] = (dur_df["Количество"] / (total_d if total_d else 1)).round(4)

                prev_cum = dur_df["Доля"].cumsum().shift(1, fill_value=0)
                dur_df["Разница с предыдущим (от 1)"] = (1 - prev_cum).clip(lower=0).round(4)

                W = dur_df["W (верх, мин)"].to_numpy(dtype=float)
                y = dur_df["Разница с предыдущим (от 1)"].to_numpy(dtype=float)
                E_fit, F_fit, G_fit = fit_bw_params(W, y)
                y_pred = bw(W, E_fit, F_fit, G_fit)

                dur_df["BW (модель)"] = np.round(y_pred, 4)
                dur_df["|ошибка|"] = np.round(np.abs(dur_df["BW (модель)"] - dur_df["Разница с предыдущим (от 1)"]), 4)

                stats = calc_stats(dur_df["Разница с предыдущим (от 1)"], dur_df["BW (модель)"])

                # Пишем таблицу
                dur_df.to_excel(writer, sheet_name=sheet, index=False, startrow=current_row)

                # Параметры/метрики — справа от таблицы (J/K)
                params_df = pd.DataFrame({
                    "Параметр": ["E", "F", "G", "MAE", "RMSE", "R2"],
                    "Значение": [
                        round(E_fit, 6),
                        round(F_fit, 6),
                        round(G_fit, 6),
                        None if stats["MAE"] is None else round(stats["MAE"], 6),
                        None if stats["RMSE"] is None else round(stats["RMSE"], 6),
                        None if stats["R2"] is None else round(stats["R2"], 6),
                    ]
                })
                params_df.to_excel(writer, sheet_name=sheet, index=False, startrow=current_row, startcol=9)

                # Сдвигаем current_row на высоту секции (заголовок + таблица + отступ)
                current_row = current_row + (MAX_MINUTES // BIN_MINUTES) + 4

    # ========== графики + формулы BW + статистика AD ==========
    wb = load_workbook(OUTPUT_WB)

    # Для статистики нужен исходный df по паре (он в памяти)
    for sh in wb.sheetnames:
        if sh == "Сводка":
            continue
        ws = wb[sh]

        # 1) график по часам
        chart_h = BarChart()
        chart_h.type = "col"
        chart_h.grouping = "clustered"
        chart_h.title = "Распределение по часам (передвижения)"
        chart_h.x_axis.title = "Час суток"
        chart_h.y_axis.title = "Передвижения"
        data_h = Reference(ws, min_col=2, min_row=1, max_row=25)
        cats_h = Reference(ws, min_col=1, min_row=2, max_row=25)
        chart_h.add_data(data_h, titles_from_data=True)
        chart_h.set_categories(cats_h)
        apply_value_only_datalabels(chart_h)
        chart_h.height = 9
        chart_h.width = 22
        ws.add_chart(chart_h, "E2")

        # 2) пройти все секции длительности (ищем строку с заголовком "Интервал (мин)")
        # и для каждой секции:
        # - проставить формулы BW/ошибка со ссылками на K (E/F/G)
        # - построить графики (абс длительность и сравнение)
        row = 1
        section_idx = 0
        while row <= ws.max_row:
            if ws.cell(row, 1).value == "Интервал (мин)":
                section_idx += 1
                dur_header_row = row
                rows_n = MAX_MINUTES // BIN_MINUTES
                dur_first = dur_header_row + 1
                dur_last = dur_header_row + rows_n

                # параметры: J/K табличка "Параметр/Значение" стоит от dur_header_row
                # строки:
                # dur_header_row: заголовки
                # dur_header_row+1: E
                # +2: F
                # +3: G
                row_E = dur_header_row + 1
                row_F = dur_header_row + 2
                row_G = dur_header_row + 3
                cell_E = f"$K${row_E}"
                cell_F = f"$K${row_F}"
                cell_G = f"$K${row_G}"

                col_W = 2
                col_diff = 5
                col_bw = 6
                col_err = 7

                for rr in range(dur_first, dur_last + 1):
                    W_cell = ws.cell(rr, col_W).coordinate
                    diff_cell = ws.cell(rr, col_diff).coordinate
                    bw_cell = ws.cell(rr, col_bw).coordinate
                    err_cell = ws.cell(rr, col_err).coordinate
                    ws[bw_cell].value = f"=1-1/POWER(1+{W_cell},{cell_E}/(1+EXP({cell_F}-{cell_G}*{W_cell})))"
                    ws[err_cell].value = f"=ABS({bw_cell}-{diff_cell})"

                # График: длительность (абс.)
                chart_d = BarChart()
                chart_d.type = "col"
                chart_d.grouping = "clustered"
                chart_d.title = f"Длительность (абс.) — секция {section_idx}"
                chart_d.x_axis.title = "Интервал, мин"
                chart_d.y_axis.title = "Передвижения"
                data_d = Reference(ws, min_col=3, min_row=dur_header_row, max_row=dur_last)  # Количество
                cats_d = Reference(ws, min_col=1, min_row=dur_first, max_row=dur_last)        # Интервалы
                chart_d.add_data(data_d, titles_from_data=True)
                chart_d.set_categories(cats_d)
                apply_value_only_datalabels(chart_d)
                chart_d.height = 8
                chart_d.width = 20
                ws.add_chart(chart_d, f"E{dur_header_row}")

                # График: сравнение Разница vs BW
                chart_cmp = LineChart()
                chart_cmp.title = f"Разница vs BW — секция {section_idx}"
                chart_cmp.x_axis.title = "W (верх), мин"
                chart_cmp.y_axis.title = "Значение"
                cats = Reference(ws, min_col=2, min_row=dur_first, max_row=dur_last)  # W
                ser_actual = Reference(ws, min_col=5, min_row=dur_header_row, max_row=dur_last)
                ser_model = Reference(ws, min_col=6, min_row=dur_header_row, max_row=dur_last)
                chart_cmp.add_data(ser_actual, titles_from_data=True)
                chart_cmp.add_data(ser_model, titles_from_data=True)
                chart_cmp.set_categories(cats)
                apply_value_only_datalabels(chart_cmp)
                chart_cmp.height = 8
                chart_cmp.width = 26
                ws.add_chart(chart_cmp, f"M{dur_header_row}")

                row = dur_last + 1
                continue
            row += 1

        # 3) статистика справа (AD)
        # определить пару по имени листа
        pair_name = sh
        sub_df = df[df["_od_pair"].astype(str).str[:31] == pair_name].copy()
        if sub_df.empty:
            sub_df = df[df["_od_pair"].astype(str).str.startswith(pair_name)].copy()

        write_stats_block(ws, start_row=1, start_col=STATS_START_COL, sub_df=sub_df, transport_group_col="_transport_group")

        for col in range(STATS_START_COL, STATS_START_COL + 8):
            ws.column_dimensions[get_column_letter(col)].width = 42

    wb.save(OUTPUT_WB)
    print(f"OK: saved -> {OUTPUT_WB}")


if __name__ == "__main__":
    main()
