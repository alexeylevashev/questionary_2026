import re
import pandas as pd
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


# Все файлы в корне
INPUT_WB = None  # будет выбран пользователем из списка .xlsx в текущей папке
INPUT_PAIRS = "Пары.xlsx"
INPUT_TRANSPORT = "Транспорт.xlsx"
OUTPUT_WB = None  # будет сформирован на основе выбранного файла


# -------------------- нормализация
def choose_input_xlsx(exclude_files=None) -> str:
    """
    Показывает список всех .xlsx в текущей папке и даёт выбрать номер.
    exclude_files: набор имён, которые не показывать (например справочники).
    """
    exclude_files = set(exclude_files or [])
    xlsx_files = sorted(
        [p for p in Path(".").iterdir()
         if p.is_file() and p.suffix.lower() == ".xlsx" and p.name not in exclude_files]
    )

    if not xlsx_files:
        raise FileNotFoundError("В текущей папке нет файлов .xlsx для выбора.")

    print("Доступные файлы Excel:")
    for i, f in enumerate(xlsx_files, start=1):
        print(f"{i}. {f.name}")

    while True:
        raw = input("Выберите номер файла: ").strip()
        try:
            idx = int(raw)
            if 1 <= idx <= len(xlsx_files):
                return xlsx_files[idx - 1].name
        except ValueError:
            pass
        print("❌ Неверный ввод. Введите число из списка.")


# -------------------- нормализация (НЕ режем строку, НЕ трогаем символы) --------------------
def normalize_excel_text(x):
    """
    Значение берём целиком. Нормализуем только Excel-артефакты:
    - неразрывные/узкие пробелы
    - переносы строк
    - множественные пробелы
    - trim
    - lower() для устойчивости
    """
    if pd.isna(x):
        return None
    s = str(x)

    # разные виды пробелов
    s = s.replace("\u00A0", " ")  # NBSP
    s = s.replace("\u2007", " ")  # figure space
    s = s.replace("\u202F", " ")  # narrow NBSP

    # переносы строк
    s = s.replace("\r", " ").replace("\n", " ")

    # схлопываем пробелы/табы
    s = re.sub(r"[ \t]+", " ", s).strip()

    return s.lower() if s else None


def is_blank(x) -> bool:
    if pd.isna(x):
        return True
    return str(x).strip() == ""


def first_transport_value(x):
    """
    Если в 'Транспорт' указано несколько систем — берём ПЕРВЫЙ тип.
    Разделители только для транспорта: , ; / перенос строки.
    """
    if pd.isna(x):
        return None
    s = str(x).strip()
    if not s:
        return None
    parts = re.split(r"[;,/\n]+", s)
    first = parts[0].strip() if parts else s.strip()
    return first if first else None


# -------------------- парсинг таблиц "столбцы = группы" --------------------
def parse_columns_as_groups(path_xlsx: str):
    """
    Формат:
      - первая строка: названия групп (заголовки столбцов)
      - ниже в каждом столбце: варианты значений, относящиеся к группе
    Возвращает:
      mapping: normalized_value -> group_name
      conflicts: normalized_value -> [groupA, groupB,...] если значение встречается в нескольких группах
    """
    df = pd.read_excel(path_xlsx, header=0, dtype=str)
    df = df.loc[:, ~df.columns.astype(str).str.lower().str.startswith("unnamed")]

    mapping = {}
    conflicts = defaultdict(list)

    for col in df.columns:
        group_name = str(col).strip()
        if not group_name:
            continue

        for raw_val in df[col].dropna().tolist():
            key = normalize_excel_text(raw_val)
            if not key:
                continue

            if key in mapping and mapping[key] != group_name:
                if mapping[key] not in conflicts[key]:
                    conflicts[key].append(mapping[key])
                if group_name not in conflicts[key]:
                    conflicts[key].append(group_name)
                continue

            mapping[key] = group_name

    return mapping, conflicts


# -------------------- Excel helpers --------------------
def style_table_header(ws, row, col_start, col_end):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def set_borders_range(ws, r1, c1, r2, c2):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = border


def write_df(ws, df, start_row, start_col, number_formats=None):
    r = start_row
    c = start_col

    for j, name in enumerate(df.columns):
        ws.cell(r, c + j, value=name)
    style_table_header(ws, r, c, c + len(df.columns) - 1)
    r += 1

    for i in range(len(df)):
        for j, col in enumerate(df.columns):
            ws.cell(r + i, c + j, value=df.iloc[i, j])

    end_row = r + len(df) - 1
    end_col = c + len(df.columns) - 1

    set_borders_range(ws, start_row, start_col, end_row, end_col)
    for rr in range(start_row, end_row + 1):
        for cc in range(start_col, end_col + 1):
            ws.cell(rr, cc).alignment = Alignment(vertical="center")

    if number_formats:
        for j0, fmt in number_formats.items():
            col = start_col + j0
            for rr in range(start_row + 1, end_row + 1):
                ws.cell(rr, col).number_format = fmt

    return end_row, end_col


def apply_value_only_datalabels(chart):
    """
    Подписи только значениями (без имён категорий/серий).
    """
    dl = DataLabelList()
    dl.showVal = True
    dl.showCatName = False
    dl.showSerName = False
    dl.showLegendKey = False
    dl.showPercent = False
    chart.dataLabels = dl


def add_column_chart_single_series(ws, title, categories_col, values_col, header_row, last_row, anchor):
    """
    Столбчатая диаграмма для одной серии (например 'Количество').
    """
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title

    data = Reference(ws, min_col=values_col, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=categories_col, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    apply_value_only_datalabels(chart)

    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, anchor)


def add_column_chart_matrix(ws, title, min_row, min_col, max_row, max_col, anchor):
    """
    Делает clustered column chart для матрицы:
      - верхняя строка: заголовки столбцов (серии)
      - первый столбец: заголовки строк (категории)
    """
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title

    data = Reference(ws, min_col=min_col + 1, min_row=min_row, max_col=max_col, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    apply_value_only_datalabels(chart)

    chart.height = 10
    chart.width = 22
    ws.add_chart(chart, anchor)


# -------------------- сопоставление пунктов --------------------
def map_od_groups(df: pd.DataFrame, pairs_map: dict):
    df = df.copy()

    df["_o_key"] = df["Пункт отправления"].apply(normalize_excel_text)
    df["_d_key"] = df["Пункт прибытия"].apply(normalize_excel_text)

    df["_o_grp"] = df["_o_key"].map(pairs_map)
    df["_d_grp"] = df["_d_key"].map(pairs_map)

    not_found_o = df.loc[df["_o_grp"].isna() & df["Пункт отправления"].notna(), "Пункт отправления"].unique()
    not_found_d = df.loc[df["_d_grp"].isna() & df["Пункт прибытия"].notna(), "Пункт прибытия"].unique()

    df["_o_grp"] = df["_o_grp"].fillna("Не определено")
    df["_d_grp"] = df["_d_grp"].fillna("Не определено")
    df["_od_pair"] = df["_o_grp"].astype(str) + " - " + df["_d_grp"].astype(str)

    return df, not_found_o, not_found_d


# -------------------- main --------------------
def main():
    global INPUT_WB, OUTPUT_WB

    # Выбор входной книги с анкетами (.xlsx) из текущей папки
    # Справочники Пары.xlsx и Транспорт.xlsx исключаем из списка, чтобы не выбрать их случайно.
    INPUT_WB = choose_input_xlsx(exclude_files={INPUT_PAIRS, INPUT_TRANSPORT})
    # Выходной файл рядом: <имя_входного>_OD.xlsx
    OUTPUT_WB = f"{Path(INPUT_WB).stem}_OD.xlsx"

    pairs_map, pairs_conflicts = parse_columns_as_groups(INPUT_PAIRS)
    transport_map, transport_conflicts = parse_columns_as_groups(INPUT_TRANSPORT)

    if pairs_conflicts:
        print("⚠ Конфликты в Пары.xlsx (одно значение в нескольких группах). Примеры:")
        for k, gs in list(pairs_conflicts.items())[:20]:
            print(f"  {k!r} -> {gs}")

    if transport_conflicts:
        print("⚠ Конфликты в Транспорт.xlsx (одно значение в нескольких группах). Примеры:")
        for k, gs in list(transport_conflicts.items())[:20]:
            print(f"  {k!r} -> {gs}")

    wb = load_workbook(INPUT_WB)
    sheetnames = wb.sheetnames
    if len(sheetnames) < 2:
        raise ValueError("Ожидались вкладки статусов, начиная со 2-й (после 'Статистика').")

    unmatched_rows = []

    for i in range(1, len(sheetnames)):
        status_sheet = sheetnames[i]
        df = pd.read_excel(INPUT_WB, sheet_name=status_sheet)

        for col in ["ID", "Пункт отправления", "Пункт прибытия", "Транспорт"]:
            if col not in df.columns:
                raise ValueError(f"На вкладке '{status_sheet}' не найден столбец '{col}'")

        total_moves = len(df)
        total_resp = int(df["ID"].nunique())

        # O/D группы
        df, nf_o, nf_d = map_od_groups(df, pairs_map)
        for v in nf_o:
            unmatched_rows.append([status_sheet, "Пункт отправления", v])
        for v in nf_d:
            unmatched_rows.append([status_sheet, "Пункт прибытия", v])

        # распределение по парам
        od_tab = df["_od_pair"].value_counts().rename_axis("Пара").reset_index(name="Количество")
        od_tab["Доля"] = od_tab["Количество"] / (total_moves if total_moves else 1)
        od_tab["На 1 респондента"] = od_tab["Количество"] / (total_resp if total_resp else 1)

        # матрицы O×D
        mat_abs = pd.pivot_table(df, index="_o_grp", columns="_d_grp", values="ID", aggfunc="size", fill_value=0)
        mat_per_resp = mat_abs / (total_resp if total_resp else 1)

        # --- ТРАНСПОРТ ---
        # Правило: пустой транспорт => Пешеход
        # Иначе: берём первый тип, маппим по справочнику, если не нашли => Не определено
        df["_t_grp"] = None

        blank_mask = df["Транспорт"].apply(is_blank)
        df.loc[blank_mask, "_t_grp"] = "Пешеход"

        non_blank = df.loc[~blank_mask, "Транспорт"].apply(first_transport_value).apply(normalize_excel_text)
        df.loc[~blank_mask, "_t_grp"] = non_blank.map(transport_map).fillna("Не определено").values

        tr_tab = df["_t_grp"].value_counts().rename_axis("Транспорт (группа)").reset_index(name="Количество")
        tr_tab["Доля"] = tr_tab["Количество"] / (total_moves if total_moves else 1)
        tr_tab["На 1 респондента"] = tr_tab["Количество"] / (total_resp if total_resp else 1)

        # --- новый лист рядом со статусом ---
        new_title = (status_sheet[:24] + "_Пары") if len(status_sheet) > 24 else (status_sheet + "_Пары")
        base = new_title
        k = 1
        while new_title in wb.sheetnames:
            new_title = (base[:28] + f"_{k}")[:31]
            k += 1

        ws_new = wb.create_sheet(title=new_title)
        idx_status = wb.sheetnames.index(status_sheet)
        wb._sheets.remove(ws_new)
        wb._sheets.insert(idx_status + 1, ws_new)

        ws_new.column_dimensions["A"].width = 44
        ws_new.column_dimensions["B"].width = 18
        ws_new.column_dimensions["C"].width = 14
        ws_new.column_dimensions["D"].width = 20
        ws_new.column_dimensions["G"].width = 42

        ws_new["A1"] = f"Статус: {status_sheet} — Пары и транспорт"
        ws_new["A1"].font = Font(size=14, bold=True)

        ws_new["A3"] = "Количество передвижений"
        ws_new["B3"] = total_moves
        ws_new["A4"] = "Количество респондентов (уник. ID)"
        ws_new["B4"] = total_resp

        # --- OD таблица ---
        ws_new["A6"] = "Распределение по парам (Группа отправления - Группа прибытия)"
        ws_new["A6"].font = Font(bold=True)

        od_start = 7
        od_end_row, _ = write_df(
            ws_new,
            od_tab[["Пара", "Количество", "Доля", "На 1 респондента"]],
            od_start, 1,
            number_formats={2: "0.00%", 3: "0.000"}
        )
        add_column_chart_single_series(ws_new, "Пары (абс.)", 1, 2, od_start, od_end_row, "G7")

        # --- матрица abs (и столбчатый график) ---
        r = od_end_row + 2
        ws_new[f"A{r}"] = "Матрица O×D (абсолютное количество)"
        ws_new[f"A{r}"].font = Font(bold=True)
        r += 1

        mat_out = mat_abs.copy()
        mat_out.insert(0, "Откуда\\Куда", mat_out.index)
        mat_out.reset_index(drop=True, inplace=True)

        mat1_start = r
        mat1_end_row, mat1_end_col = write_df(ws_new, mat_out, mat1_start, 1)
        add_column_chart_matrix(ws_new, "O×D (абс.)", mat1_start, 1, mat1_end_row, mat1_end_col, f"G{mat1_start}")

        # --- матрица per respondent (и столбчатый график) ---
        r = mat1_end_row + 2
        ws_new[f"A{r}"] = "Матрица O×D (кол-во / число респондентов статуса)"
        ws_new[f"A{r}"].font = Font(bold=True)
        r += 1

        mat2_out = mat_per_resp.copy()
        mat2_out.insert(0, "Откуда\\Куда", mat2_out.index)
        mat2_out.reset_index(drop=True, inplace=True)

        mat2_start = r
        mat2_end_row, mat2_end_col = write_df(
            ws_new, mat2_out, mat2_start, 1,
            number_formats={j: "0.000" for j in range(1, len(mat2_out.columns))}
        )
        add_column_chart_matrix(ws_new, "O×D (на респондента)", mat2_start, 1, mat2_end_row, mat2_end_col, f"G{mat2_start}")

        # --- транспорт общий ---
        r = mat2_end_row + 2
        ws_new[f"A{r}"] = "Распределение по транспорту (группы)"
        ws_new[f"A{r}"].font = Font(bold=True)
        r += 1

        tr_start = r
        tr_end_row, _ = write_df(
            ws_new,
            tr_tab[["Транспорт (группа)", "Количество", "Доля", "На 1 респондента"]],
            tr_start, 1,
            number_formats={2: "0.00%", 3: "0.000"}
        )
        add_column_chart_single_series(ws_new, "Транспорт (абс.)", 1, 2, tr_start, tr_end_row, f"G{tr_start}")

        # --- транспорт по каждой OD паре ---
        r = tr_end_row + 2
        ws_new[f"A{r}"] = "Транспорт по каждой паре (O-D)"
        ws_new[f"A{r}"].font = Font(bold=True)
        r += 1

        for od_name in od_tab["Пара"].tolist():
            sub = df[df["_od_pair"] == od_name]
            if sub.empty:
                continue

            moves_od = len(sub)
            sub_tr = sub["_t_grp"].value_counts().rename_axis("Транспорт (группа)").reset_index(name="Количество")
            sub_tr["Доля"] = sub_tr["Количество"] / (moves_od if moves_od else 1)
            sub_tr["На 1 респондента"] = sub_tr["Количество"] / (total_resp if total_resp else 1)

            ws_new[f"A{r}"] = f"{od_name} (передвижений: {moves_od})"
            ws_new[f"A{r}"].font = Font(bold=True)
            r += 1

            block_start = r
            block_end_row, _ = write_df(
                ws_new,
                sub_tr[["Транспорт (группа)", "Количество", "Доля", "На 1 респондента"]],
                block_start, 1,
                number_formats={2: "0.00%", 3: "0.000"}
            )
            add_column_chart_single_series(ws_new, f"Транспорт для пары: {od_name}", 1, 2, block_start, block_end_row, f"G{block_start}")
            r = block_end_row + 2

        ws_new.freeze_panes = "A7"

    # Лист несопоставленных
    if "НЕ_СОПОСТАВЛЕНО" in wb.sheetnames:
        wb.remove(wb["НЕ_СОПОСТАВЛЕНО"])

    ws_u = wb.create_sheet("НЕ_СОПОСТАВЛЕНО")
    ws_u.column_dimensions["A"].width = 26
    ws_u.column_dimensions["B"].width = 22
    ws_u.column_dimensions["C"].width = 80

    ws_u["A1"] = "Статус"
    ws_u["B1"] = "Поле"
    ws_u["C1"] = "Значение (как в анкетах)"
    style_table_header(ws_u, 1, 1, 3)

    for idx, (st, fld, val) in enumerate(unmatched_rows, start=2):
        ws_u.cell(idx, 1, st)
        ws_u.cell(idx, 2, fld)
        ws_u.cell(idx, 3, val)

    if unmatched_rows:
        set_borders_range(ws_u, 1, 1, 1 + len(unmatched_rows), 3)
    ws_u.freeze_panes = "A2"

    wb.save(OUTPUT_WB)
    print(f"OK: saved -> {OUTPUT_WB}")


if __name__ == "__main__":
    main()
