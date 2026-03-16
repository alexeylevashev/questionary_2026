import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


def choose_input_xlsx(exclude_names=None) -> str:
    """
    Дает пользователю выбрать входной .xlsx файл из текущей папки (корня проекта).
    exclude_names: список/набор имён файлов, которые не показывать в списке.
    """
    exclude_names = set(exclude_names or [])
    files = sorted(
        [p for p in Path(".").iterdir()
         if p.is_file() and p.suffix.lower() == ".xlsx" and p.name not in exclude_names]
    )

    if not files:
        raise FileNotFoundError("В текущем каталоге нет .xlsx файлов для выбора.")

    # Если файл ровно один — выбираем автоматически
    if len(files) == 1:
        print(f"Найден один .xlsx файл: {files[0].name}")
        return files[0].name

    print("Доступные .xlsx файлы:")
    for i, f in enumerate(files, start=1):
        print(f"{i}. {f.name}")

    while True:
        raw = input("Выберите номер файла: ").strip()
        try:
            idx = int(raw)
            if 1 <= idx <= len(files):
                return files[idx - 1].name
        except ValueError:
            pass
        print("❌ Неверный ввод. Введите число из списка.")


def build_status_to_group_map(soc: pd.DataFrame) -> dict:
    status_to_group = {}
    for group in soc.columns:
        values = soc[group].dropna().astype(str).str.strip()
        for v in values:
            if v:
                status_to_group[v] = str(group).strip()
    return status_to_group


def add_value_only_labels(chart: BarChart):
    dl = DataLabelList()
    dl.showVal = True
    dl.showCatName = False
    dl.showSerName = False
    dl.showPercent = False
    dl.showLegendKey = False
    chart.dataLabels = dl


def create_chart(ws, col_idx, title, pos):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = ""
    chart.x_axis.title = ""

    data = Reference(ws, min_col=col_idx, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # подписи данных: только значения
    add_value_only_labels(chart)

    ws.add_chart(chart, pos)


def main():
    # Пользователь выбирает анкетный файл из текущего каталога
    input_file = choose_input_xlsx(exclude_names={"Соцгруппы.xlsx"})
    if not Path("Соцгруппы.xlsx").exists():
        raise FileNotFoundError("Не найден файл 'Соцгруппы.xlsx' в текущем каталоге.")

    print(f"\nИспользуется файл анкет: {input_file}")

    # -------------------------------
    # 1. Загрузка данных
    # -------------------------------
    df = pd.read_excel(input_file)
    soc = pd.read_excel("Соцгруппы.xlsx")

    # -------------------------------
    # 2. Формирование словаря: статус -> соцгруппа
    # -------------------------------
    status_to_group = build_status_to_group_map(soc)

    # -------------------------------
    # 3. Назначение соцгруппы
    # -------------------------------
    if "Социальный статус" not in df.columns:
        raise KeyError("В анкетах не найден столбец 'Социальный статус'.")
    if "ID" not in df.columns:
        raise KeyError("В анкетах не найден столбец 'ID'.")

    df["Соцгруппа"] = df["Социальный статус"].astype(str).str.strip().map(status_to_group)
    df["Соцгруппа"] = df["Соцгруппа"].fillna("Не определено")

    # -------------------------------
    # 4. Статистика по группам
    # -------------------------------
    stats = (
        df.groupby("Соцгруппа")
          .agg(
              Передвижения=("ID", "count"),
              Респонденты=("ID", pd.Series.nunique)
          )
          .reset_index()
    )

    stats["Среднее передвижений"] = (stats["Передвижения"] / stats["Респонденты"]).round(2)

    # -------------------------------
    # 5. Запись в Excel
    # -------------------------------
    stem = Path(input_file).stem
    output_file = f"{stem}_Статус.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        stats.to_excel(writer, sheet_name="Статистика", index=False)

        for group, group_df in df.groupby("Соцгруппа"):
            sheet_name = str(group)[:31]  # ограничение Excel
            group_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # -------------------------------
    # 6. Добавление графиков
    # -------------------------------
    wb = load_workbook(output_file)
    ws = wb["Статистика"]

    create_chart(ws, 2, "Количество передвижений", "G2")
    create_chart(ws, 3, "Количество респондентов", "G20")
    create_chart(ws, 4, "Среднее передвижений на респондента", "G38")

    # -------------------------------
    # 7. Сохранение
    # -------------------------------
    wb.save(output_file)
    print(f"Готово: файл '{output_file}' создан.")


if __name__ == "__main__":
    main()
